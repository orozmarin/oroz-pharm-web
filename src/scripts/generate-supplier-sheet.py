"""
Generira Excel tablicu za unos artikala po dobavljaču.
Struktura: Kategorija | Podkategorija | Naziv artikla (10 praznih redova po podkategoriji)

Pokrenuti s: python3 src/scripts/generate-supplier-sheet.py
Izlaz: ~/Desktop/Dobavljac.xlsx
"""

import os

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Instaliraj openpyxl: pip3 install openpyxl")
    exit(1)

# ── Podaci ────────────────────────────────────────────────────────────────────

categories = [
    ("Sjemenski program", [
        "Ratarskih kultura",
        "Uljarice",
        "Trave i djeteline",
        "Hibridi povrća",
        "Ostalo",
    ]),
    ("Gnojiva", [
        "Mineralna",
        "Organska",
        "Biostimulatori",
        "Poboljšivači tla",
        "Ostalo",
    ]),
    ("Zaštita bilja", [
        "Herbicidi",
        "Fungicidi",
        "Insekticidi",
        "Limacidi",
        "Biocidi/Rodenticidi",
        "Ostalo",
    ]),
    ("Stočna hrana", [
        "Sano",
        "TSH Čakovec",
        "Patent",
        "Genera",
        "Lek",
        "Valpovka",
        "Ostalo",
    ]),
    ("Sadni materijal", [
        "Voćne sadnice",
        "Sadnice vinove loze",
        "Prijesadnice cvijeća",
        "Lučica",
        "Sjemenski krumpir",
        "Ostalo",
    ]),
    ("Oprema za povrtlarstvo", [
        "Klasmann",
        "Florabella",
        "Brill",
        "Hawita",
        "Plantella",
        "Plastenička folija",
        "Mreže",
        "Ostalo",
    ]),
    ("Navodnjavanje", [
        "Pumpe",
        "Spojnice i ventili",
        "Crijeva kap po kap",
        "Rasprskivači",
        "Ostalo",
    ]),
    ("Enologija", [
        "Boce",
        "Kvasci",
        "Inox bačve",
        "Kanisteri",
        "Ostalo",
    ]),
    ("Oprema za pčelarstvo", [
        "Staklenke",
        "Poklopci",
        "Pogače",
        "Košnice",
        "Ostalo",
    ]),
    ("Poljoprivredni strojevi", [
        "Kosilice",
        "Motorne pile",
        "Trimeri",
        "Ostalo",
    ]),
    ("Alati", [
        "Škare i pile za rezidbu",
        "Alatke i držala",
    ]),
    ("Oprema za stočarstvo", [
        "Hranilice i pojilice",
        "Pastiri i oprema",
    ]),
    ("Ulja i maziva", [
        "Ulja",
        "Maziva",
        "Antifriz",
        "Tekućine za stakla",
    ]),
    ("Kućni ljubimci", [
        "Hrana za golubove",
        "Hrana i oprema za pse i mačke",
        "Hrana za ptice i ribe",
    ]),
    ("Pribor za kolinje", [
        "Crijeva",
        "Punilice",
        "Mesoreznice",
        "Noževi",
        "Kuke",
    ]),
    ("Vrt i okućnica", [
        "Tačke",
        "Nit za košnju",
        "Rukavice",
        "Čizme",
        "Lampioni",
    ]),
    ("Roba široke potrošnje", [
        "Kace",
        "Bačve",
        "Kante",
        "Lončanice za cvijeće",
        "Sprejevi",
    ]),
    ("Gume i zračnice", []),
    ("Ekološki proizvodi", []),
]

ROWS_PER_SUB = 10  # praznih redova po podkategoriji

# ── Stilovi ───────────────────────────────────────────────────────────────────

GREEN_DARK   = "1A5C2E"
GREEN_LIGHT  = "D6EAD9"
GREY_HEADER  = "3D3D3D"
WHITE        = "FFFFFF"

def header_font():
    return Font(name="Calibri", bold=True, color=WHITE, size=11)

def cat_font():
    return Font(name="Calibri", bold=True, color=WHITE, size=10)

def sub_font():
    return Font(name="Calibri", bold=True, color=GREEN_DARK, size=10)

def data_font():
    return Font(name="Calibri", size=10)

def thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def cat_fill():
    return PatternFill("solid", fgColor=GREEN_DARK)

def sub_fill():
    return PatternFill("solid", fgColor=GREEN_LIGHT)

def header_fill():
    return PatternFill("solid", fgColor=GREY_HEADER)

def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def left():
    return Alignment(horizontal="left", vertical="center", wrap_text=False)

# ── Workbook ──────────────────────────────────────────────────────────────────

wb = Workbook()
ws = wb.active
ws.title = "Dobavljač"

# Zamrzni gornji red
ws.freeze_panes = "A2"

# Širine stupaca
ws.column_dimensions["A"].width = 28
ws.column_dimensions["B"].width = 28
ws.column_dimensions["C"].width = 50

# ── Zaglavlje ─────────────────────────────────────────────────────────────────

headers = ["Kategorija", "Podkategorija", "Naziv artikla"]
for col, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.font = header_font()
    cell.fill = header_fill()
    cell.alignment = center()
    cell.border = thin_border()

ws.row_dimensions[1].height = 22

# ── Podaci ────────────────────────────────────────────────────────────────────

current_row = 2

for cat_name, subs in categories:
    if not subs:
        # Kategorija bez podkategorija — jedan blok
        row_start = current_row
        for _ in range(ROWS_PER_SUB):
            ws.cell(row=current_row, column=1, value=cat_name).font = cat_font()
            ws.cell(row=current_row, column=1).fill = cat_fill()
            ws.cell(row=current_row, column=1).alignment = left()
            ws.cell(row=current_row, column=1).border = thin_border()
            ws.cell(row=current_row, column=2).border = thin_border()
            ws.cell(row=current_row, column=3).font = data_font()
            ws.cell(row=current_row, column=3).border = thin_border()
            ws.cell(row=current_row, column=3).alignment = left()
            ws.row_dimensions[current_row].height = 18
            current_row += 1
        # Spoji ćelije kategorije
        ws.merge_cells(f"A{row_start}:A{current_row - 1}")
        ws.cell(row=row_start, column=1).alignment = center()
    else:
        cat_start = current_row
        for sub_name in subs:
            row_start = current_row
            for i in range(ROWS_PER_SUB):
                # Kategorija (samo prvi red bloka)
                cat_cell = ws.cell(row=current_row, column=1)
                cat_cell.value = cat_name if i == 0 else None
                cat_cell.fill = cat_fill()
                cat_cell.alignment = left()
                cat_cell.border = thin_border()

                # Podkategorija
                sub_cell = ws.cell(row=current_row, column=2)
                sub_cell.value = sub_name if i == 0 else None
                sub_cell.fill = sub_fill()
                sub_cell.font = sub_font()
                sub_cell.alignment = left()
                sub_cell.border = thin_border()

                # Naziv artikla
                art_cell = ws.cell(row=current_row, column=3)
                art_cell.font = data_font()
                art_cell.border = thin_border()
                art_cell.alignment = left()

                ws.row_dimensions[current_row].height = 18
                current_row += 1

            # Spoji ćelije podkategorije
            ws.merge_cells(f"B{row_start}:B{current_row - 1}")
            ws.cell(row=row_start, column=2).alignment = center()

        # Spoji ćelije kategorije
        ws.merge_cells(f"A{cat_start}:A{current_row - 1}")
        ws.cell(row=cat_start, column=1).alignment = center()
        ws.cell(row=cat_start, column=1).font = cat_font()
        ws.cell(row=cat_start, column=1).fill = cat_fill()

# ── Spremi ────────────────────────────────────────────────────────────────────

output_path = os.path.expanduser("~/Desktop/Dobavljac.xlsx")
wb.save(output_path)
print(f"Dokument spremljen: {output_path}")
